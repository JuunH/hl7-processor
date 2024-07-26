# This file is licensed under the GNU General Public License v3.0


from openpyxl import Workbook
import json


class HL7:
    """This class creates an object to represent a HL7 message.

    It contains in-built functions to structure the HL7 in different forms to
    provide an easier way for viewing and manipulation.

    This can be used in any HL7 related development.

    The Embedded list, Dictionary and JSON is already generated in the object for you.

    Usage:

        import hl7

        with open("hl7_message.txt", "r") as file:
            message_content = file.read()
            hl7_message = hl7.HL7(message_content)
    """
    def __init__(self, hl7_message) -> None:
        self.hl7_message = hl7_message.strip()
        self._segments = []
        self.elist = None
        self.dict = None
        self.json = None
        self.hl7_to_list(hl7_message)
        self.list_to_dict()
        self.list_to_json()


    def hl7_to_list(self, hl7_message=None):
        """Transforms a HL7 message to an embedded list.

        Assumes within the message that each segment is already separated by 
        new lines.
            
        Any fields with subfields will create another list while single value 
        fields will not and instead be stored as a regular string.

        e.g. ['MSH', ... ['ADT', 'A28'], [[1, 2, 3], [4, 5, 6]]...]

        Args:
            message (string): The HL7 message

        Returns:
            list: Stores HL7 message as an embedded list within the object
        """
        data = []

        if not hl7_message:
            hl7_message = self.hl7_message

        for segment in hl7_message.splitlines():
            fields = segment.split('|')
            self._segments.append(fields.pop(0)) # Retrieve segment ID and append to separate list.
            segment_data = []

            for field in fields:
                if field == '^~\&':
                    # Prevents splitting the encoding characters
                    segment_data.extend(['|', '^~\&'])
                    continue
                if '~' in field:
                    # Process repeating fields
                    repetition_data = [repetition.split('^') for repetition in field.split('~')]
                    segment_data.append(repetition_data)
                else:
                    if len(field.split('^')) == 1:
                        # Process fields
                        segment_data.extend(field.split('^'))    
                    else:
                        # Process subfields
                        segment_data.append(field.split('^'))

            data.append(segment_data)
        self._sort_segments()

        self.elist = data


    def list_to_hl7(self):
        """Rebuilds the HL7 message from its embedded list structure.

        IMPORTANT: changes made to Dict form will not change the JSON form and 
        vice versa, so use the right form before exporting the raw HL7, or copy
        the changes over.

        e.g:

            hl7.dict['MSH.1'] = hl7.json['MSH'][1]
            # or 
            hl7.json['MSH'][1] = hl7.dict['MSH.1']

        Returns:
            str: Stores the rebuilt HL7 message within the object
        """
        self.hl7 = self._gen_hl7(self.elist)
    

    def list_to_dict(self):
        """Constructs a dictionary of the HL7 message from the embedded list structure.

        Enables to reference any part of a HL7.

        e.g.

        hl7['MSH.1']
        hl7['EVN.5.6']
        hl7['PID.3.1.5']

        Returns: 
            dict: Stores the HL7 as a dictionary within the object.
        """
        self.dict = self._gen_dict(self.elist)


    def dict_to_list(self):
        """Transforms the dictionary form to an embedded list structure.

        It utilises the size of the key/reference (e.g MSH.1 is size 1) to
        know how to construct each field

        Args:
            dictionary (dict): Dictionary form of HL7

        Returns:
            list: Stores resulting transformation within the object.
        """
        segments = {}
        subfield_list = []
        repeats = []
        repeats_list = []
        prev_segment = ''
        prev_depth = 1
        prev_subfield_position = 1
        prev_repeat_position = 1
        count = 1 
        dict_size = len(self.dict)

        for key, value in self.dict.items():
            segment_id = key.split('.')[0]
            # Determines whether if its a field, subfield, or repeat
            current_depth = len(key.split('.'))-1
            
            if segment_id not in segments:
                segments[segment_id] = []
            
            if current_depth == prev_depth:
                match current_depth:
                    case 1:
                        # Field
                        segments[segment_id].append(value)

                    case 2:
                        # Subfields
                        if prev_subfield_position < int(key.split('.')[1]):
                            segments[segment_id].append(subfield_list)
                            subfield_list = []
                            prev_subfield_position = int(key.split('.')[1])
                        subfield_list.append(value)

                    case 3:
                        # Repeat Fields
                        if prev_repeat_position < int(key.split('.')[2]):
                            repeats_list.append(repeats)
                            repeats = []
                            prev_repeat_position = int(key.split('.')[2])
                        repeats.append(value)
                        
            elif current_depth > prev_depth:
                match current_depth:
                    case 2: 
                        # Subfields
                        subfield_list.append(value)
                        prev_subfield_position = int(key.split('.')[1])
                    case 3:
                        # Repeat fields
                        if subfield_list:
                            segments[segment_id].append(subfield_list)
                            subfield_list = []
                        repeats.append(value)
                        prev_repeat_position = int(key.split('.')[2])

            elif current_depth < prev_depth:
                match current_depth:
                    case 1:
                        if repeats_list:
                            # Fields
                            repeats_list.append(repeats) if repeats else None
                            segments[segment_id].append(repeats_list)
                            repeats_list = []
                            repeats = []
                            prev_repeat_position = 1
                        
                        if subfield_list:
                            if segment_id != prev_segment:
                                segments[prev_segment].append(subfield_list)
                            else:
                                segments[segment_id].append(subfield_list)
                            subfield_list = []
                            prev_subfield_position = 1

                        segments[segment_id].append(value)

                    case 2:
                        # Subfields
                        repeats_list.append(repeats) if repeats else None
                        segments[segment_id].append(repeats_list)
                        subfield_list.append(value)
                        prev_subfield_position = int(key.split('.')[1])
                        repeats_list = []
                        repeats = []
                        prev_repeat_position = 1
            
            if count == dict_size:
                # Stores any remaining data once reached end of dictionary
                if repeats_list:
                    repeats_list.append(repeats) if repeats else None
                    segments[segment_id].append(repeats_list)
                    repeats_list = []
                    repeats = []
                    prev_repeat_position = 1
                        
                if subfield_list:
                    if segment_id != prev_segment:
                        segments[prev_segment].append(subfield_list)
                    else:
                        segments[segment_id].append(subfield_list)
                    subfield_list = []
                    prev_subfield_position = 1

            count += 1
            prev_segment = segment_id
            prev_depth = current_depth
        
        result = []

        for key, value in segments.items():
            result.append(value)
        
        self.elist = result
    

    def list_to_json(self):
        """Constructs a dictionary for JSON output and for any code that requires
        iterating over any fields. Can be seen as an alternate dictionary structure
        for a HL7 message/

        Referencing a field is similar to the normal dictionary form.

        e.g. hl7['PID.7.5'] is the same as:

            json_hl7['PID'][7][5]
        """
        segment_pointer = 0
        dictionary = dict()

        for item in self.elist:
            segment = self._segments[segment_pointer]
            dictionary[segment] = []
            dictionary[segment].extend(self._gen_json(item))
            segment_pointer += 1

        self.json = dictionary
        

    def json_to_list(self):
        """Transforms the JSON form to an embedded list structure

        Returns:
            list: Stores embedded list structure within the object
        """
        elist = []
        for key in self.json:
            elist.append(self.json[key]) 
        remove_fillers = lambda elist: [
            item if isinstance(item, str) else remove_fillers(item[1:]) 
            for item in elist
        ]
        elist = remove_fillers(elist)
        return elist


    def structure(self):
        """Constructs a visual text output of the HL7 structure along with its reference.

        Returns:
            str: Formatted representation of the HL7 structure.
        """
        return self._gen_structure(self.elist)
    

    def export(self, format, filepath):
        """Export the HL7 message to a chosen format.

        Args:
            format (string): format of either txt, excel, or json
            filepath (string): location to save message output
        """
        match format:
            case 'Excel' | 'excel' | 'xlsx':
                wb = Workbook()
                ws = wb.active
                row = 1
                for key, value in self.dict.items():
                    ws.cell(row=row, column=1).value = key
                    ws.cell(row=row, column=2).value = value
                    row += 1
                wb.save(f"{filepath}{'.xlsx' if not filepath.endswith('.xlsx') else ''}")

            case 'dict' | 'dictionary':
                with open(f'{filepath}.txt', 'w') as file:
                    file.write(str(self.dict))
            
            case 'json' | 'JSON':
                with open(f"{filepath}{'.json' if not filepath.endswith('.json') else ''}", 'w') as file:
                    json.dump(self.json, file, indent=4)

            case 'Structure' | 'structure':
                with open(f"{filepath}{'.txt' if not filepath.endswith('.txt') else ''}", 'w') as file:
                    file.write(self.structure())
 
            case 'list' | 'List':
                with open(f"{filepath}{'.txt' if not filepath.endswith('.txt') else ''}", 'w') as file:
                    file.write(str(self.elist))

            case 'raw' | 'Raw':
                with open(f"{filepath}{'.txt' if not filepath.endswith('.txt') else ''}", 'w') as file:
                    file.write(self.hl7)
    

    def _gen_hl7(self, elist, depth=0, next_segment=False):
        """Rescursive function to rebuild the HL7 message for list_to_hl7()

        Args:
            elist (_type_): List representation of HL7 message
            depth (int, optional): Current depth within the list. Defaults to 0.

        Returns:
            _type_: string of original HL7 message
        """
        message = ''
        separators = ('', '\n', '|', '^', '~')
        seg_pointer = 0

        for item in elist:
            if isinstance(item, list):
                if not next_segment:
                    segment = self._segments[seg_pointer].split('_')[0]
                    message += f"{segment}|"
                    message += self._gen_hl7(item, depth + 1, True)
                    seg_pointer += 1 if seg_pointer < len(self._segments)-1 else 0
                else:
                    message += self._gen_hl7(item, depth + 1, True)
            else:
                if item == '|':
                    continue 
                message += f'{item}{separators[depth+1] if depth < 3 else separators[depth]}'
        message.strip()
        return f'{message[:-1]}{separators[depth] if depth < 3 else separators[depth+1]}'
    

    def _gen_dict(self, data, path=''):
        """Recursive function to transform an embedded list to dictionary for
        list_to_dict()

        Args:
            data (list): HL7 message as a list
            path (str): The current path within the structure. Defaults to ''.

        Returns:
            dictionary: HL7 message with index as the key
        """

        hl7_dictionary = dict()
        if isinstance(data, list):
            for index, item in enumerate(data):
                new_path = f'{path}.{index+1}' if path else str(self._segments[index])
                hl7_dictionary.update(self._gen_dict(item, new_path))
            return hl7_dictionary
        else:
            return {path: data}
    

    def _gen_json(self, data):
        """Recursive function to transform an embedded list to JSON form for
        list_to_json()"""
        if isinstance(data, list):
            segment_data = ['']
            for item in data:
                segment_data.append(self._gen_json(item))
            return segment_data
        else:
            return data
    

    def _gen_structure(self, data, path = '', level=0, indent = -8):
        """Constructs a visual HL7 structure along with its field references.

        Args:
            data (list): HL7 message as a list 
            path (str): The current path within the structure. Defaults to ''.
            indent (int): The indentation level for formatting. Defaults to -8.

        Returns:
            str: The formatted HL7 structure representation.
        """
        result = ''
        padding = ' ' * indent if indent >= 0 else ''
        if isinstance(data, list):
            for index, item in enumerate(data):
                new_path = f'{path}.{index+1}' if path else str(self._segments[index])
                result += str(self._gen_structure(item, new_path, level+1, indent+4))
        else:
            return f'{padding}{path}: {data}\n' if data else f'{padding}{path}: *empty*\n'
        return result
    

    def _sort_segments(self):
        """Sorts out any duplicate segment IDs stored from repeat fields or segments 
        after conversion

        For example, a repeat in the PID will be renamed to PID_1. This is to ensures
        the other functions are run properly to ensure repeats are recorded properly
        and not overwriting the first repeat
        
        Returns:
            list: list of segments retrieved from the original message.
        """
        seen = {}
        sorted_segments = []
        for segment in self._segments:
            if segment in seen:
                seen[segment] += 1
                new_item = f"{segment}_{seen[segment]}"
            else:
                seen[segment] = 0
                new_item = segment
            sorted_segments.append(new_item)
        self._segments = sorted_segments


def comparison(original, result, filepath):
    """Compares the data in each field position from two HL7 messages. Outputs
    the result in an Excel sheet with a written formula to compare the outputs. 

    Args:
        original (dictionary): Dictionary structure of a HL7 message
        result (dictionary): Dictionary structure of a transformed HL7 message
        filepath (string): Filepath to desired save location of the Excel result
    """
    wb = Workbook()
    ws = wb.active
    i = 1
    for key, value in original.dict.items():
        ws[f'A{i}'].value = key
        ws[f'B{i}'].value = value
        ws[f'C{i}'].value = result.dict[key] if key in result.dict else None
        ws[f'D{i}'].value = f'=IF(B{i}=C{i}, "MATCH", "NO MATCH")'
        i += 1
    wb.save(f"{filepath}{'.xlsx' if not filepath.endswith('.xlsx') else ''}")
